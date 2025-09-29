import openpyxl
from openpyxl.styles import PatternFill, Alignment

def export_to_excel(results, filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "So khop"

    # Header
    headers = ["STT", "Nội dung", "Phiếu chuyển", "Hợp đồng", "Tỉ lệ so khớp", "Mức độ"]
    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")

    # Colors
    color_map = {
        "XANH": "90EE90",  # xanh nhạt
        "VÀNG": "FFD700",  # vàng
        "ĐỎ": "FF6347"     # đỏ
    }

    # Data rows
    for idx, row in enumerate(results, start=1):
        ws.append([
            idx,
            row["truong"],
            row["phieu_chuyen"],
            row["hop_dong"],
            f'{row["ty_le"]}%',
            row["mau"]
        ])

        # tô màu theo mức độ
        fill = PatternFill(start_color=color_map[row["mau"]],
                           end_color=color_map[row["mau"]],
                           fill_type="solid")
        ws.cell(idx+1, 6).fill = fill

    wb.save(filename)
